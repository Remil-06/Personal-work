import os
import random
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill

# Lista de nombres específicos para las columnas de "Mesa"
nombres_mesa = ["ADD NAMES","ADD NAMES","ETC.."] #Escibe los nombres de la gente

# Lista de nombres para las columnas de Plataforma, Micros 1 y Micros 2
nombres_plataforma = ["ADD NAMES","ADD NAMES","ETC.."] #Escibe los nombres de la gente

# Nombres especiales para fechas específicas
nombres_fecha_especial = {}


# Función para equilibrar la asignación basada en el total de apariciones
def equilibrar_asignacion_total(nombres_disponibles, conteo_total):
    # Encontrar el mínimo conteo actual
    min_conteo = min(conteo_total[nombre] for nombre in nombres_disponibles)
    # Filtrar los nombres que tienen el conteo más bajo
    nombres_equilibrados = [nombre for nombre in nombres_disponibles if conteo_total[nombre] == min_conteo]
    # Elegir un nombre de entre los que tienen menos apariciones
    return random.choice(nombres_equilibrados)


# Función para generar la tabla
def generar_tabla():
    # Crear un rango de fechas 
    fechas = []
    fecha_inicio = datetime(2025, 1, 1)
    fecha_fin = datetime(2025, 3, 31)
    while fecha_inicio <= fecha_fin:
        if fecha_inicio.weekday() in [1, 6]:  # Martes y domingos
            fechas.append(fecha_inicio.strftime('%d/%m/%Y'))
        fecha_inicio += timedelta(days=1)

    mesas = []
    plataformas = []
    micros_1_list = []
    micros_2_list = []
    
    nombres_previos = {
        "Mesa": [],
        "Plataforma": [],
        "Micros 1": [],
        "Micros 2": []
    }

    # Conteo de asignaciones por columna y total para cada persona
    conteo_columna = {
        "Mesa": {nombre: 0 for nombre in nombres_mesa + nombres_plataforma},
        "Plataforma": {nombre: 0 for nombre in nombres_mesa + nombres_plataforma},
        "Micros 1": {nombre: 0 for nombre in nombres_mesa + nombres_plataforma},
        "Micros 2": {nombre: 0 for nombre in nombres_mesa + nombres_plataforma}
    }
    
    conteo_total = {nombre: 0 for nombre in nombres_mesa + nombres_plataforma}
    
    for fecha in fechas:
        nombres_especiales = nombres_fecha_especial.get(fecha, [])
        
        # Elegir nombres para Mesa
        mesa_disponibles = [nombre for nombre in nombres_mesa if nombre not in nombres_especiales and nombre not in nombres_previos["Mesa"][-2:]]
        if not mesa_disponibles:
            mesa_disponibles = [nombre for nombre in nombres_mesa if nombre not in nombres_especiales]
        mesa = equilibrar_asignacion_total(mesa_disponibles, conteo_total)
        mesas.append(mesa)
        nombres_previos["Mesa"].append(mesa)
        conteo_columna["Mesa"][mesa] += 1
        conteo_total[mesa] += 1
        
        # Elegir nombres para Plataforma, Micros 1 y Micros 2
        plataforma_disponibles = [nombre for nombre in nombres_plataforma if nombre not in [mesa] + nombres_especiales and nombre not in nombres_previos["Plataforma"][-2:]]
        if not plataforma_disponibles:
            plataforma_disponibles = [nombre for nombre in nombres_plataforma if nombre not in [mesa] + nombres_especiales]
        plataforma = equilibrar_asignacion_total(plataforma_disponibles, conteo_total)
        plataformas.append(plataforma)
        nombres_previos["Plataforma"].append(plataforma)
        conteo_columna["Plataforma"][plataforma] += 1
        conteo_total[plataforma] += 1
        
        micros_1_disponibles = [nombre for nombre in nombres_plataforma if nombre not in [mesa, plataforma] + nombres_especiales and nombre not in nombres_previos["Micros 1"][-2:]]
        if not micros_1_disponibles:
            micros_1_disponibles = [nombre for nombre in nombres_plataforma if nombre not in [mesa, plataforma] + nombres_especiales]
        micros_1 = equilibrar_asignacion_total(micros_1_disponibles, conteo_total)
        micros_1_list.append(micros_1)
        nombres_previos["Micros 1"].append(micros_1)
        conteo_columna["Micros 1"][micros_1] += 1
        conteo_total[micros_1] += 1
        
        micros_2_disponibles = [nombre for nombre in nombres_plataforma if nombre not in [mesa, plataforma, micros_1] + nombres_especiales and nombre not in nombres_previos["Micros 2"][-2:]]
        if not micros_2_disponibles:
            micros_2_disponibles = [nombre for nombre in nombres_plataforma if nombre not in [mesa, plataforma, micros_1] + nombres_especiales]
        micros_2 = equilibrar_asignacion_total(micros_2_disponibles, conteo_total)
        micros_2_list.append(micros_2)
        nombres_previos["Micros 2"].append(micros_2)
        conteo_columna["Micros 2"][micros_2] += 1
        conteo_total[micros_2] += 1
    
    # Crear un DataFrame de pandas con los datos generados
    data = {
        'Fecha': fechas,
        'Mesa': mesas,
        'Plataforma': plataformas,
        'Micros 1': micros_1_list,
        'Micros 2': micros_2_list
    }
    tabla = pd.DataFrame(data)
    
    return tabla, conteo_columna, conteo_total

# Generar la tabla, conteos por columna y conteos totales
tabla, conteo_columna, conteo_total = generar_tabla()


print(tabla)

# Guardar la tabla en un archivo Excel
output_file = 'c:\\Users\\06rem\\OneDrive\\Desktop\\Listado_de_sonido.xlsx'
tabla.to_excel(output_file, index=False, engine='openpyxl')

# Aplicar estilos condicionales
wb = openpyxl.load_workbook(output_file)
ws = wb.active

fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    fecha = datetime.strptime(row[0].value, '%d/%m/%Y')
    if fecha.weekday() == 1:  # Martes
        for cell in row:
            cell.fill = fill

wb.save(output_file)
print(f"Tabla guardada en {output_file}")

# Crear un DataFrame con los conteos por columna y el total
conteos_df = pd.DataFrame({
    'Mesa': conteo_columna['Mesa'],
    'Plataforma': conteo_columna['Plataforma'],
    'Micros 1': conteo_columna['Micros 1'],
    'Micros 2': conteo_columna['Micros 2'],
    'Total': conteo_total
}).fillna(0).astype(int)

# Imprimir el DataFrame de conteos
print(conteos_df)

# Guardar el DataFrame de conteos en un archivo Excel
conteos_output_file = 'c:\\Users\\06rem\\OneDrive\\Desktop\\Conteos_de_sonido.xlsx'
conteos_df.to_excel(conteos_output_file, index=True, engine='openpyxl')
print(f"Conteos guardados en {conteos_output_file}")
